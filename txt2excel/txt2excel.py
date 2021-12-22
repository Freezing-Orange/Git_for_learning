from openpyxl import Workbook,load_workbook
from openpyxl.styles import *
import warnings
import re

def txtToList(fileName):
    with open(fileName,'r') as f:
        res=f.readlines()
    return res

def handleList(txtList):
    res=[]
    i=0
    length=len(txtList)
    while i<length:
        if "TABLE" in txtList[i] and "CHECK#" in txtList[i]:
            temp=re.findall(r"[0-9]+", txtList[i])
            tableNo, checkNo=temp[0],temp[1]
            i+=1
            while True:
                if "Cashier" in txtList[i]:
                    CashierNo=re.findall(r"[0-9]+", txtList[i])[0]
                    break
                else:
                    i+=1
            res.append([tableNo,checkNo,CashierNo])
    return res

def createOpenExcel(Name="Summary.xlsx"):
    wb1=Workbook()
    ws1=wb.create_sheet("sheet1")
    return ws1

warnings.filterwarnings('ignore')